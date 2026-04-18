# src/utils/export.py
# Handles Excel export and BibTeX generation.
# Called from main.py via Streamlit download buttons.

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict


def export_to_excel(papers: List[Dict], clusters: Dict, query: str) -> bytes:
    """
    Builds an Excel file with two sheets:
      Sheet 1 — all papers sorted by cluster, with cluster name column
      Sheet 2 — cluster summary (theme, count, top paper, year range)

    Returns bytes object suitable for st.download_button.
    """
    # Build lookup: paper title -> cluster number and name
    paper_to_cluster_no   = {}
    paper_to_cluster_name = {}
    for cid, info in clusters.items():
        for p in info.get('papers', []):
            key = p.get('title', '')[:50].lower().strip()
            paper_to_cluster_no[key]   = int(cid) + 1
            paper_to_cluster_name[key] = info.get('name', f'Cluster {int(cid)+1}')

    # Build main rows
    rows = []
    for p in papers:
        key = p.get('title', '')[:50].lower().strip()
        authors = p.get('authors', [])
        if isinstance(authors, list):
            authors_str = '; '.join(str(a) for a in authors[:3])
            if len(authors) > 3:
                authors_str += ' et al.'
        else:
            authors_str = str(authors)

        # Extract AI summary fields safely — works whether LLM ran or not
        _summary   = p.get('ai_summary') or {}
        _problem   = _summary.get('Research_Problem', '') or ''
        _findings  = _summary.get('Key_Findings', []) or []
        _lit_para  = _summary.get('Literature_Review_Paragraph', '') or ''

        # Key findings: join list into numbered lines for readability in Excel
        if isinstance(_findings, list):
            _findings_str = '\n'.join(
                f"{i+1}. {f}" for i, f in enumerate(_findings) if f
            )
        else:
            _findings_str = str(_findings)

    
        rows.append({
            'Cluster No.':         paper_to_cluster_no.get(key, ''),
            'Cluster Theme':       paper_to_cluster_name.get(key, ''),
            'Title':               p.get('title', ''),
            'Authors':             authors_str,
            'Year':                p.get('year', ''),
            'Source':              p.get('source', ''),
            'Citations':           p.get('citations', 'N/A'),
            'Relevance Score':     p.get('relevance_score', ''),
            'Abstract':            p.get('abstract', '')[:300],
            'Problem Statement':   _problem[:500],
            'Key Findings':        _findings_str[:600],
            'Literature Review':   _lit_para[:800],
            'URL':                 p.get('url', '')
        })

    df = pd.DataFrame(rows)
    # Sort by cluster then by citations descending
    if not df.empty and 'Cluster No.' in df.columns:
        df = df.sort_values(
            ['Cluster No.', 'Citations'],
            ascending=[True, False],
            na_position='last'
        )

    # Build summary rows
    summary_rows = []
    for cid, info in clusters.items():
        cp    = info.get('papers', [])
        years = [int(p['year']) for p in cp
                 if str(p.get('year', '')).isdigit()]
        top   = max(cp, key=lambda p: p.get('citations', 0) or 0) \
                if cp else {}
        summary_rows.append({
            'Cluster No.':   int(cid) + 1,
            'Theme':         info.get('name', ''),
            'Paper Count':   len(cp),
            'Top Paper':     top.get('title', '')[:60],
            'Top Citations': top.get('citations', 0) or 0,
            'Year Range':    f"{min(years)}–{max(years)}" if years else ''
        })
    summary_df = pd.DataFrame(summary_rows)

    cluster_colors = [
        'FFFFFF', 'EBF5FB', 'FEF9E7', 'EAFAF1',
        'FDEDEC', 'F4ECF7', 'FDF2E9', 'E8F8F5'
    ]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Papers')
        summary_df.to_excel(writer, index=False, sheet_name='Cluster Summary')

        # ── Style: Papers sheet ──────────────────────────
        ws = writer.sheets['Papers']

        # Bold blue header
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='2E75B6')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Column widths
        col_widths = {
            'A': 10, 'B': 28, 'C': 52, 'D': 28,
            'E': 8,  'F': 18, 'G': 10, 'H': 14,
            'I': 45, 'J': 52, 'K': 55, 'L': 65, 'M': 45
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # Colour rows by cluster number
        for row in ws.iter_rows(min_row=2):
            cnum = row[0].value
            if cnum and isinstance(cnum, int):
                color = cluster_colors[(cnum - 1) % len(cluster_colors)]
                for cell in row:
                    cell.fill      = PatternFill('solid', fgColor=color)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.freeze_panes = 'A2'

        # ── Style: Summary sheet ─────────────────────────
        ws2 = writer.sheets['Cluster Summary']
        for cell in ws2[1]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='2E75B6')
            cell.alignment = Alignment(horizontal='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2.column_dimensions[col].width = 18
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['D'].width = 55

    return output.getvalue()


def generate_bibtex(papers: List[Dict]) -> str:
    """
    Generates a BibTeX string for all papers.
    User downloads as .bib and drops into Overleaf.
    """
    entries = []
    seen_keys: set = set()

    for p in papers:
        authors  = p.get('authors', [])
        if isinstance(authors, list):
            author_str = ' and '.join(str(a) for a in authors)
            surname    = str(authors[0]).split()[-1].lower() \
                         if authors else 'unknown'
        else:
            author_str = str(authors)
            surname    = author_str.split()[-1].lower() \
                         if author_str.strip() else 'unknown'

        year  = str(p.get('year', '0000'))
        title = p.get('title', 'Untitled')
        url   = p.get('url', '')

        # Ensure unique BibTeX key
        base_key = f"{surname}{year}"
        key      = base_key
        suffix   = 1
        while key in seen_keys:
            key    = f"{base_key}{chr(96 + suffix)}"   # a, b, c …
            suffix += 1
        seen_keys.add(key)

        entry = (
            f"@article{{{key},\n"
            f"  title   = {{{title}}},\n"
            f"  author  = {{{author_str}}},\n"
            f"  year    = {{{year}}},\n"
            f"  url     = {{{url}}}\n"
            f"}}"
        )
        entries.append(entry)

    return '\n\n'.join(entries)

